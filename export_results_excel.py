from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nss_paths.topology import adjacency_summary


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "outputs"
SUMMARY_CSV = OUTPUT_DIR / "wyniki_podsumowanie.csv"
DETAILS_CSV = OUTPUT_DIR / "wyniki_szczegolowe.csv"
OUTPUT_XLSX = OUTPUT_DIR / "porownanie_algorytmow_trasy_rozlaczne.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="1F4E79", bold=True, size=14)


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofilter_and_freeze(ws, last_col: int, last_row: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "A2"


def write_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    style_header(ws[1])


def build_differences(details: pd.DataFrame) -> pd.DataFrame:
    pairs = {}
    for row in details.itertuples(index=False):
        key = (row.zrodlo, row.cel)
        pairs.setdefault(key, {})[row.algorytm] = row

    rows = []
    for (source, target), values in pairs.items():
        naive = values.get("naiwny")
        suurballe = values.get("suurballe")
        if naive is None or suurballe is None:
            continue
        diff = suurballe.laczna_liczba_laczy - naive.laczna_liczba_laczy
        if diff != 0:
            rows.append(
                {
                    "Para": f"{source}-{target}",
                    "naiwny P1": naive.trasa_podstawowa,
                    "naiwny P2": naive.trasa_zabezpieczajaca,
                    "Suurballe P1": suurballe.trasa_podstawowa,
                    "Suurballe P2": suurballe.trasa_zabezpieczajaca,
                    "różnica sumy": diff,
                }
            )
    return pd.DataFrame(rows)


def build_workbook() -> Path:
    summary = pd.read_csv(SUMMARY_CSV)
    details = pd.read_csv(DETAILS_CSV)

    summary = summary.rename(
        columns={
            "liczba_zadan": "żądania",
            "zaakceptowane": "ok",
            "odrzucone": "odrz.",
            "wspolczynnik_odrzucenia": "R",
            "srednia_dlugosc_podstawowa": "śr. P1",
            "srednia_dlugosc_zabezpieczajaca": "śr. P2",
            "srednia_laczna_liczba_laczy": "śr. suma",
        }
    )
    details = details.rename(
        columns={
            "zrodlo": "s",
            "cel": "t",
            "zaakceptowane": "ok",
            "trasa_podstawowa": "P1",
            "trasa_zabezpieczajaca": "P2",
            "dlugosc_podstawowa": "len P1",
            "dlugosc_zabezpieczajaca": "len P2",
            "laczna_liczba_laczy": "suma",
            "powod_odrzucenia": "powód",
        }
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Podsumowanie"

    ws["A1"] = "Porównanie algorytmów doboru tras rozłącznych"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:H1")
    ws.append([])
    ws.append(["Liczba węzłów", 21])
    ws.append(["Liczba łączy", 35])
    ws.append(["Graf spójny", "tak"])
    ws.append(["Spójność krawędziowa", 2])
    ws.append([])
    start_row = ws.max_row + 1
    ws.append(list(summary.columns))
    for row in summary.itertuples(index=False):
        ws.append(list(row))
    style_header(ws[start_row])
    for row in ws.iter_rows(min_row=start_row + 1, max_row=ws.max_row, min_col=5, max_col=8):
        row[0].number_format = "0.00%"
        for cell in row[1:]:
            cell.number_format = "0.00"

    ws.append([])
    ws.append(["Różnica Suurballe - naiwny", "Wartość"])
    style_header(ws[ws.max_row])
    naive = summary[summary["algorytm"] == "naiwny"].iloc[0]
    suurballe = summary[summary["algorytm"] == "suurballe"].iloc[0]
    delta_rows = [
        ("Zmiana odrzuceń", suurballe["odrz."] - naive["odrz."]),
        ("Zmiana współczynnika odrzucenia", suurballe["R"] - naive["R"]),
        ("Zmiana średniej długości podstawowej", suurballe["śr. P1"] - naive["śr. P1"]),
        ("Zmiana średniej długości zabezpieczającej", suurballe["śr. P2"] - naive["śr. P2"]),
        ("Zmiana średniej łącznej liczby łączy", suurballe["śr. suma"] - naive["śr. suma"]),
    ]
    for label, value in delta_rows:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 13

    ws_details = wb.create_sheet("Szczegoly")
    write_dataframe(ws_details, details)
    widths = [12, 8, 8, 8, 32, 36, 10, 10, 10, 18]
    for index, width in enumerate(widths, start=1):
        ws_details.column_dimensions[get_column_letter(index)].width = width
    autofilter_and_freeze(ws_details, len(details.columns), len(details) + 1)

    ws_diff = wb.create_sheet("Roznice")
    differences = build_differences(pd.read_csv(DETAILS_CSV))
    write_dataframe(ws_diff, differences)
    for index, width in enumerate([12, 30, 34, 30, 34, 14], start=1):
        ws_diff.column_dimensions[get_column_letter(index)].width = width
    autofilter_and_freeze(ws_diff, len(differences.columns), len(differences) + 1)

    ws_topology = wb.create_sheet("Topologia")
    ws_topology["A1"] = "Lista sąsiedztwa topologii A4"
    ws_topology["A1"].font = TITLE_FONT
    ws_topology["A1"].fill = TITLE_FILL
    ws_topology.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws_topology["A2"] = "Sąsiedzi to węzły połączone bezpośrednio łączem o koszcie 1."
    ws_topology.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws_topology.append(["Węzeł", "Sąsiedzi", "Stopień"])
    for row in adjacency_summary():
        ws_topology.append(list(row))
    style_header(ws_topology[3])
    ws_topology.column_dimensions["A"].width = 10
    ws_topology.column_dimensions["B"].width = 34
    ws_topology.column_dimensions["C"].width = 10
    ws_topology.freeze_panes = "A4"

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


if __name__ == "__main__":
    print(build_workbook())
